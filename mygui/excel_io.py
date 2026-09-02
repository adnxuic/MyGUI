"""Import Excel workbooks and preview their sheets before creating tables."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, cast
import zipfile

import openpyxl

from PySide6.QtCore import Qt
from PySide6.QtGui import QBrush, QColor
from PySide6.QtWidgets import (
    QCheckBox,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFormLayout,
    QHBoxLayout,
    QHeaderView,
    QLineEdit,
    QMessageBox,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from mygui.application_theme import bind_widget_qss, current_qss_tokens, watch_qss_tokens
from mygui.widgets.english_buttons import apply_english_dialog_buttons
from mygui.database import ColumnRef, ColumnType, SheetDocument, TableChangeSet, TableMutationCommand
from mygui.database.table_document import DEFAULT_ROWS, infer_column_type, new_id, validate_component_name
from mygui.resource_limits import load_resource_limits


EXCEL_WORKBOOK_SUFFIXES = frozenset({".xlsx", ".xlsm"})
EXCEL_FILE_FILTER = "Excel Workbooks (*.xlsx *.xlsm)"
EXCEL_PREVIEW_ROWS = 8


def is_supported_excel_workbook(file_name: str | Path) -> bool:
    """Return whether supported excel workbook."""

    path = Path(file_name)
    return path.is_file() and path.suffix.casefold() in EXCEL_WORKBOOK_SUFFIXES


def validate_excel_workbook(file_name: str | Path) -> Path:
    """Validate excel workbook."""

    path = Path(file_name)
    if path.suffix.casefold() not in EXCEL_WORKBOOK_SUFFIXES:
        raise ValueError("Only .xlsx and .xlsm Excel workbooks are supported.")
    if not path.is_file():
        raise ValueError(f"Excel workbook does not exist: {path}")
    limits = load_resource_limits()
    if path.stat().st_size > limits.max_excel_bytes:
        raise ValueError("Excel workbook exceeds the configured byte budget.")
    try:
        with zipfile.ZipFile(path) as archive:
            uncompressed_size = sum(info.file_size for info in archive.infolist())
    except (OSError, zipfile.BadZipFile) as exc:
        raise ValueError("Excel workbook is not a valid Office ZIP container.") from exc
    if uncompressed_size > limits.max_excel_uncompressed_bytes:
        raise ValueError(
            "Excel workbook exceeds the configured uncompressed byte budget."
        )
    return path


@dataclass
class ExcelSheetData:
    """Represent the application's excel sheet data."""

    name: str
    rows: list[list[Any]]


@dataclass
class ExcelColumnSpec:
    """Describe excel column spec values shared across application layers."""

    name: str
    type: ColumnType
    values: list[Any]


@dataclass
class ExcelSheetSpec:
    """Describe excel sheet spec values shared across application layers."""

    source_name: str
    target_name: str
    columns: list[ExcelColumnSpec]


def read_excel_workbook(file_name: str) -> list[ExcelSheetData]:
    """Read excel workbook."""

    path = validate_excel_workbook(file_name)
    limits = load_resource_limits()
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if len(workbook.worksheets) > limits.max_excel_sheets:
            raise ValueError("Excel workbook exceeds the configured sheet budget.")
        result = []
        cell_count = 0
        for worksheet in workbook.worksheets:
            declared_cells = int(worksheet.max_row) * int(worksheet.max_column)
            if declared_cells > limits.max_excel_cells - cell_count:
                raise ValueError("Excel workbook exceeds the configured cell budget.")
            rows = []
            for row in worksheet.iter_rows(values_only=True):
                values = list(row)
                cell_count += len(values)
                if cell_count > limits.max_excel_cells:
                    raise ValueError(
                        "Excel workbook exceeds the configured cell budget."
                    )
                rows.append(values)
            result.append(ExcelSheetData(worksheet.title, rows))
        return result
    finally:
        workbook.close()


def _column_values(rows: list[list[Any]], column: int, skip_header: bool) -> list[Any]:
    start = 1 if skip_header else 0
    return [row[column] if column < len(row) else None for row in rows[start:]]


def _default_column_name(rows: list[list[Any]], column: int, use_header: bool) -> str:
    if use_header and rows and column < len(rows[0]) and rows[0][column] not in (None, ""):
        return str(rows[0][column]).strip()
    return f"Column {column + 1}"


class ExcelSheetPreview(QWidget):
    """Provide the excel sheet preview Qt widget."""

    INCLUDE_ROW = 0
    NAME_ROW = 1
    TYPE_ROW = 2
    SAMPLE_START_ROW = 3

    def __init__(self, sheet: ExcelSheetData, parent=None, use_header: bool = True):
        super().__init__(parent)
        self.sheet = sheet
        self.include = QCheckBox("Import this sheet", self)
        self.include.setChecked(True)
        self.header = QCheckBox("Use first row as column names", self)
        self.header.setChecked(use_header)
        self.target_name = QLineEdit(sheet.name, self)
        self.columns = QTableWidget(self)
        self._column_include_boxes: list[QCheckBox] = []
        self.columns.setAlternatingRowColors(True)
        self.columns.setCornerButtonEnabled(False)
        horizontal_header = self.columns.horizontalHeader()
        horizontal_header.setSectionResizeMode(QHeaderView.Interactive)
        horizontal_header.setDefaultSectionSize(190)
        horizontal_header.setMinimumSectionSize(110)
        horizontal_header.setStretchLastSection(False)
        self.columns.verticalHeader().setDefaultSectionSize(28)

        form = QFormLayout()
        form.addRow("Target sheet:", self.target_name)
        layout = QVBoxLayout(self)
        layout.addWidget(self.include)
        layout.addWidget(self.header)
        layout.addLayout(form)
        layout.addWidget(self.columns)
        self.header.toggled.connect(self.rebuild)
        self.rebuild()
        watch_qss_tokens(self, lambda _tokens: self._refresh_sample_chrome())

    def _chrome_color(self, token: str) -> QColor:
        return QColor(current_qss_tokens()[token])

    def _refresh_sample_chrome(self) -> None:
        for column in range(self.columns.columnCount()):
            if column < len(self._column_include_boxes):
                self._set_column_included(
                    column,
                    self.column_include_checkbox(column).isChecked(),
                )

    def rebuild(self):
        """Rebuild the control from the latest available application state."""

        max_columns = max((len(row) for row in self.sheet.rows), default=0)
        previous_names = []
        previous_included = [checkbox.isChecked() for checkbox in self._column_include_boxes]
        for column in range(self.columns.columnCount()):
            name_editor = self.columns.cellWidget(self.NAME_ROW, column)
            previous_names.append(
                name_editor.text() if isinstance(name_editor, QLineEdit) else ""
            )

        column_values = [
            _column_values(self.sheet.rows, column, self.header.isChecked())
            for column in range(max_columns)
        ]
        sample_count = min(
            EXCEL_PREVIEW_ROWS,
            max((len(values) for values in column_values), default=0),
        )
        source_row_start = 2 if self.header.isChecked() else 1

        self.columns.clear()
        self.columns.setColumnCount(max_columns)
        self.columns.setRowCount(self.SAMPLE_START_ROW + sample_count)
        self.columns.setVerticalHeaderLabels(
            ["Import", "Column Name", "Type"]
            + [f"Source Row {source_row_start + row}" for row in range(sample_count)]
        )
        self._column_include_boxes = []

        used_names = set()
        for column in range(max_columns):
            default_name = _default_column_name(self.sheet.rows, column, self.header.isChecked())
            name = previous_names[column] if column < len(previous_names) and previous_names[column] else default_name
            base = name or f"Column {column + 1}"
            suffix = 2
            while name.casefold() in used_names:
                name = f"{base} {suffix}"
                suffix += 1
            used_names.add(name.casefold())
            values = column_values[column]
            inferred = infer_column_type(values)

            header_item = QTableWidgetItem(name)
            header_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            self.columns.setHorizontalHeaderItem(column, header_item)

            include_checkbox = QCheckBox(self.columns)
            include_checkbox.setChecked(
                previous_included[column] if column < len(previous_included) else True
            )
            include_container = QWidget(self.columns)
            include_layout = QHBoxLayout(include_container)
            include_layout.setContentsMargins(0, 0, 0, 0)
            include_layout.addWidget(include_checkbox, alignment=Qt.AlignCenter)
            self.columns.setCellWidget(self.INCLUDE_ROW, column, include_container)
            self._column_include_boxes.append(include_checkbox)

            name_editor = QLineEdit(name, self.columns)
            name_editor.setClearButtonEnabled(True)
            name_editor.textChanged.connect(
                lambda text, index=column: self._update_column_header(index, text)
            )
            self.columns.setCellWidget(self.NAME_ROW, column, name_editor)

            type_combo = QComboBox(self.columns)
            type_combo.addItems([column_type.value for column_type in ColumnType])
            type_combo.setCurrentText(inferred.value)
            self.columns.setCellWidget(self.TYPE_ROW, column, type_combo)

            for sample_index in range(sample_count):
                value = values[sample_index] if sample_index < len(values) else None
                sample_item = QTableWidgetItem("" if value is None else str(value))
                sample_item.setFlags(sample_item.flags() & ~Qt.ItemIsEditable)
                if isinstance(value, bool):
                    sample_item.setTextAlignment(Qt.AlignCenter)
                elif isinstance(value, (int, float)):
                    sample_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                else:
                    sample_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                if value is None:
                    sample_item.setBackground(self._chrome_color("COLOR_SURFACE_ALT"))
                self.columns.setItem(self.SAMPLE_START_ROW + sample_index, column, sample_item)

            self.columns.setColumnWidth(column, 190)
            include_checkbox.toggled.connect(
                lambda checked, index=column: self._set_column_included(index, checked)
            )
            self._set_column_included(column, include_checkbox.isChecked())

    def _update_column_header(self, column: int, text: str):
        item = self.columns.horizontalHeaderItem(column)
        if item is not None:
            item.setText(text.strip() or f"Column {column + 1}")

    def column_name_editor(self, column: int) -> QLineEdit:
        """Return the column name editor."""

        return cast(QLineEdit, self.columns.cellWidget(self.NAME_ROW, column))

    def column_type_editor(self, column: int) -> QComboBox:
        """Return the column type editor."""

        return cast(QComboBox, self.columns.cellWidget(self.TYPE_ROW, column))

    def column_include_checkbox(self, column: int) -> QCheckBox:
        """Return the column include checkbox."""

        return self._column_include_boxes[column]

    def _set_column_included(self, column: int, included: bool):
        self.column_name_editor(column).setEnabled(included)
        self.column_type_editor(column).setEnabled(included)
        for row in range(self.SAMPLE_START_ROW, self.columns.rowCount()):
            item = self.columns.item(row, column)
            if item is None:
                continue
            if included:
                item.setForeground(QBrush())
                item.setBackground(
                    self._chrome_color("COLOR_SURFACE_ALT") if not item.text() else QBrush()
                )
            else:
                item.setForeground(self._chrome_color("COLOR_BORDER_STRONG"))
                item.setBackground(self._chrome_color("COLOR_HOVER_LIGHT"))

    def spec(self) -> ExcelSheetSpec | None:
        """Return the current spec."""

        if not self.include.isChecked():
            return None
        target_name = validate_component_name(self.target_name.text(), "Sheet name")
        names = set()
        columns = []
        for column in range(self.columns.columnCount()):
            if not self.column_include_checkbox(column).isChecked():
                continue
            name = self.column_name_editor(column).text().strip()
            if not name:
                raise ValueError(f"Column {column + 1} in {self.sheet.name} has no name.")
            if name.casefold() in names:
                raise ValueError(f"Duplicate column name in {self.sheet.name}: {name}")
            names.add(name.casefold())
            type_combo = self.column_type_editor(column)
            values = _column_values(self.sheet.rows, column, self.header.isChecked())
            columns.append(ExcelColumnSpec(name, ColumnType(type_combo.currentText()), values))
        if not columns and self.columns.columnCount() > 0:
            raise ValueError(f"Select at least one column to import from {self.sheet.name}.")
        return ExcelSheetSpec(self.sheet.name, target_name, columns)


class ExcelImportDialog(QDialog):
    """Provide the excel import dialog Qt widget."""

    def __init__(self, sheets: list[ExcelSheetData], parent=None):
        super().__init__(parent)
        self.setWindowTitle("Import Excel")
        bind_widget_qss(
            self,
            "mygui/widgets/title_bar/titlebar_dialog/dialog_style.qss",
        )
        self.resize(960, 600)
        self.setMinimumSize(720, 480)
        self.tabs = QTabWidget(self)
        self.pages = []
        for sheet in sheets:
            page = ExcelSheetPreview(sheet, self)
            self.pages.append(page)
            self.tabs.addTab(page, sheet.name)
        buttons = apply_english_dialog_buttons(
            QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self)
        )
        buttons.accepted.connect(self._validate_and_accept)
        buttons.rejected.connect(self.reject)
        layout = QVBoxLayout(self)
        layout.addWidget(self.tabs)
        layout.addWidget(buttons)

    def _validate_and_accept(self):
        try:
            specs = self.specs()
            if not specs:
                raise ValueError("Select at least one sheet to import.")
        except ValueError as exc:
            QMessageBox.warning(self, "Import Excel", str(exc))
            return
        self.accept()

    def specs(self) -> list[ExcelSheetSpec]:
        """Return the current specs."""

        result = []
        target_names = set()
        for page in self.pages:
            spec = page.spec()
            if spec is None:
                continue
            normalized = spec.target_name.casefold()
            if normalized in target_names:
                raise ValueError(f"Duplicate target sheet name: {spec.target_name}")
            target_names.add(normalized)
            result.append(spec)
        return result


def _default_specs(sheets: list[ExcelSheetData]) -> list[ExcelSheetSpec]:
    result = []
    for sheet in sheets:
        page = ExcelSheetPreview(sheet)
        result.append(page.spec())
        page.deleteLater()
    return [spec for spec in result if spec is not None]


def import_excel_into_table(
    file_name: str,
    table,
    parent=None,
    show_preview: bool = True,
    *,
    publish_new_project: bool = True,
):
    """Import excel into table."""

    sheets = read_excel_workbook(file_name)
    if not sheets:
        raise ValueError("The workbook contains no sheets.")
    if show_preview:
        dialog = ExcelImportDialog(sheets, parent)
        if dialog.exec() != QDialog.Accepted:
            return None
        specs = dialog.specs()
    else:
        specs = _default_specs(sheets)

    return import_sheet_specs_into_table(
        file_name,
        specs,
        table,
        command_text="Import Excel",
        reason="excel-import",
        publish_new_project=publish_new_project,
    )


def import_sheet_specs_into_table(file_name: str, specs: list[ExcelSheetSpec], table,
                                  command_text: str = "Import Data",
                                  reason: str = "data-import", *,
                                  publish_new_project: bool = True):
    """Commit already validated tabular import specs as one Repository command."""

    subtable = table.current_subtable()
    created_project = subtable is None
    project = None if subtable is None else subtable.project
    used_sheet_names = {
        sheet.name.casefold() for sheet in project.sheets.values()
    } if project is not None else set()

    def unique_sheet_name(preferred: str) -> str:
        base = validate_component_name(preferred, "Sheet name")
        candidate = base
        suffix = 2
        while candidate.casefold() in used_sheet_names:
            candidate = f"{base} {suffix}"
            suffix += 1
        used_sheet_names.add(candidate.casefold())
        return candidate

    imported_sheets = []
    for spec in specs:
        target_name = unique_sheet_name(spec.target_name)
        document = SheetDocument(id=new_id(), name=target_name, row_count=max(
            DEFAULT_ROWS, max((len(column.values) for column in spec.columns), default=0)
        ))
        for column in spec.columns:
            document.add_column(column.name, column.type, values=column.values)
        if not document.columns:
            document.add_column("Column 1")
        imported_sheets.append(document)

    if subtable is None:
        project_name = validate_component_name(Path(file_name).stem, "Project name")
        subtable = table.create_project_table(
            project_name,
            publish=False,
        )
        project = subtable.project
    placeholder_sheets = list(project.sheets.values()) if created_project else []

    def redo():
        if created_project:
            subtable._dispose_tabs()
            for sheet in placeholder_sheets:
                project.sheets.pop(sheet.id, None)
        for sheet in imported_sheets:
            project.sheets[sheet.id] = sheet
        subtable._build_tabs()

    def undo():
        subtable._dispose_tabs()
        for sheet in imported_sheets:
            project.sheets.pop(sheet.id, None)
        if created_project:
            for sheet in placeholder_sheets:
                project.sheets[sheet.id] = sheet
        subtable._build_tabs()

    def changes():
        refs = {
            ColumnRef(project.id, sheet.id, column.id)
            for sheet in imported_sheets for column in sheet.columns
        }
        return TableChangeSet(
            project.id, refs, metadata_changed=True, structure_changed=True, reason=reason
        )

    try:
        table.repository.push(project.id, TableMutationCommand(
            command_text, table.repository, project.id, redo, undo, changes
        ))
    except Exception:
        if created_project:
            table.remove_project_table(
                project.id,
                publish=False,
            )
        raise
    if created_project and publish_new_project:
        table.repository.publish_project_added(project.id)
    return subtable


def import_excel_into_workspace(file_name: str, table, figure_window=None, parent=None,
                                show_preview: bool = True):
    """Import one workbook and create its canvas when no project is active."""
    path = validate_excel_workbook(file_name)
    create_canvas = figure_window is not None and figure_window.current_canva is None
    stage_new_project = create_canvas and table.current_subtable() is None
    subtable = import_excel_into_table(
        str(path),
        table,
        parent=parent or table,
        show_preview=show_preview,
        publish_new_project=not stage_new_project,
    )
    if subtable is None:
        return None
    if create_canvas:
        try:
            width, height, dpi = figure_window.creation_figure_size()
            figure_window.add_figure(
                width=width,
                height=height,
                dpi=dpi,
                style="default",
                canva_name=subtable.project.name,
                create_table=False,
            )
        except Exception:
            if stage_new_project:
                table.remove_project_table(
                    subtable.project.id,
                    publish=False,
                )
            raise
        if stage_new_project:
            table.repository.publish_project_added(subtable.project.id)
    return subtable
