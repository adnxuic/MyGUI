from Qt_core import *

from code.widgets.table.py_table import PyTable
from code.widgets.title_bar.py_title_button import SelectMenuButton, MenuButton, StaticSelectButton, DynSelectButton, \
    PullDownButton
from code.widgets.title_bar.py_pull_down_menu import StyleMenu
from code.widgets.title_bar.titlebar_dialog.py_title_bar_dialog import PyStyleDialog, PyLayoutDialog
from code.widgets.title_bar.titlebar_dialog.py_chart_dialog import chart_dialog_dict
from code.widgets.title_bar.titlebar_dialog.py_element_dialog import element_dialog_dict

import openpyxl as xl

import json
import os

current_path = os.path.dirname(os.path.abspath(__file__))


class SelectorMenuBar(QFrame):
    def __init__(self, stacklayout_bottom=None):
        super().__init__()

        self.setObjectName("selector_menu_bar")

        self.layout = QHBoxLayout(self)
        self.layout.setContentsMargins(0, 0, 0, 0)

        self.stacklayout_bottom = stacklayout_bottom

        # 设置按钮组
        self.buttonGroup = QButtonGroup(self)
        self.buttonGroup.setExclusive(True)  # 设置互斥

        # 添加按钮
        self.style_button = SelectMenuButton('style', 'pictures/icons/style.svg')
        self.style_button.setChecked(True)
        self.style_button.toggled.connect(self.the_button_was_toggled)
        self.layout.addWidget(self.style_button)
        self.buttonGroup.addButton(self.style_button)

        self.layout_button = SelectMenuButton('layout', 'pictures/icons/layout.svg')
        self.layout_button.toggled.connect(self.the_button_was_toggled)
        self.layout.addWidget(self.layout_button)
        self.buttonGroup.addButton(self.layout_button)

        self.chart_button = SelectMenuButton('chart', 'pictures/icons/chart.svg')
        self.chart_button.toggled.connect(self.the_button_was_toggled)
        self.layout.addWidget(self.chart_button)
        self.buttonGroup.addButton(self.chart_button)

        self.element_button = SelectMenuButton('element', 'pictures/icons/element.svg')
        self.element_button.toggled.connect(self.the_button_was_toggled)
        self.layout.addWidget(self.element_button)
        self.buttonGroup.addButton(self.element_button)

    def the_button_was_toggled(self, checked):
        if not checked:
            return

        if self.style_button.isChecked():
            self.stacklayout_bottom.setCurrentIndex(0)
        elif self.layout_button.isChecked():
            self.stacklayout_bottom.setCurrentIndex(1)
        elif self.chart_button.isChecked():
            self.stacklayout_bottom.setCurrentIndex(2)
        elif self.element_button.isChecked():
            self.stacklayout_bottom.setCurrentIndex(3)


class MenuBar(QFrame):
    def __init__(self, table: PyTable):
        super().__init__()

        self.table = table

        self.setObjectName("menu_bar")
        self.layout = QHBoxLayout(self)
        self.layout.setContentsMargins(0, 0, 0, 4)

        # 创建触发文件菜单的按钮
        self.file_button = MenuButton('file', 'pictures/icons/file.svg', self)
        self.file_button.clicked.connect(lambda: self.show_menu(self.file_menu, self.file_button))
        self.layout.addWidget(self.file_button)

        # 添加分割线
        separator = QFrame(self)
        separator.setFrameShape(QFrame.VLine)  # 设置为垂直线类型
        separator.setFrameShadow(QFrame.Sunken)  # 给线一个凹陷的外观
        self.layout.addWidget(separator)

        # 创建触发编辑菜单的按钮
        self.edit_button = QPushButton('edit', self)
        self.edit_button.setObjectName('menu_button')
        self.edit_button.clicked.connect(lambda: self.show_menu(self.edit_menu, self.edit_button))
        self.layout.addWidget(self.edit_button)

        # 添加弹性空间
        spacer = QSpacerItem(10, 20, QSizePolicy.Expanding, QSizePolicy.Minimum)
        self.layout.addItem(spacer)

        # 设置文件菜单
        self.file_menu = QMenu(self)
        self.initActions()

        # 设置编辑菜单
        self.edit_menu = QMenu(self)
        self.edit_menu.addAction('copy')
        self.edit_menu.addAction('paste')
        self.edit_menu.addAction('cut')

    def show_menu(self, menu_name, button_name):
        # 显示菜单
        menu_name.exec(button_name.mapToGlobal(button_name.rect().bottomLeft()))

    def initActions(self):
        file_open_action = QAction(QIcon("pictures/icons/open.svg"), "打开", self.file_menu)
        file_open_action.triggered.connect(self.open_file)

        file_save_action = QAction(QIcon("pictures/icons/save.svg"), "保存", self.file_menu)
        file_save_action.triggered.connect(self.save_file)

        self.file_menu.addAction(file_open_action)
        self.file_menu.addAction(file_save_action)

    def open_file(self):
        # 打开Excel文件
        file_name, _ = QFileDialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx *.xls)")
        if os.path.exists(file_name):
            subtable = self.table.add_new_table(is_open=True)
            wb = xl.load_workbook(file_name)
            # 读取所有工作表
            for i, sheet_name in enumerate(wb.sheetnames):
                # 添加新工作表
                if i > 0:
                    subtable.add_new_sheet()

                tableview = subtable.get_table(i)
                sheet = wb[sheet_name]
                # 读取所有列
                for j, col in enumerate(sheet.iter_cols()):
                    data_list = [cell.value for cell in col]
                    tableview.add_excel_data(data_list, j)
                # 保存
                tableview.model.save_data_to_database()
            wb.close()

    def save_file(self):
        print('save file')


class ControlBar(QFrame):
    def __init__(self, parent=None):
        super().__init__()
        # 设置对象名称
        self.setObjectName("control_bar")
        # 设置布局
        self.layout = QHBoxLayout(self)

        self.parent = parent

        # 添加弹性空间
        spacer = QSpacerItem(10, 20, QSizePolicy.Expanding, QSizePolicy.Minimum)
        self.layout.addItem(spacer)

        # 添加最小化按钮
        minmize_button = QPushButton(QIcon("pictures/icons/minimize.svg"), "")
        minmize_button.setObjectName("minimize_button")
        minmize_button.clicked.connect(self.parent.showMinimized)
        self.layout.addWidget(minmize_button)

        # 添加最大化按钮

        # 添加关闭按钮
        button_close = QPushButton(QIcon("pictures/icons/close.svg"), "")
        button_close.setObjectName("close_button")
        button_close.clicked.connect(self.parent.close)
        self.layout.addWidget(button_close)


class SelectorStyleMenuBar(QFrame):
    def __init__(self, figure_window=None, fig_control_window=None):
        super().__init__()

        # 读取可用的样式
        style_json_path = os.path.join(current_path, 'available_styles.json')
        with open(style_json_path, 'r') as json_file:
            self.available_styles_dict = json.load(json_file)

        # 设置对象名称
        self.setObjectName("selector_menu")

        # 设置布局
        self.layout = QHBoxLayout()
        self.layout.setContentsMargins(0, 0, 0, 10)

        self.button_dict = {}

        for index, style in enumerate(self.available_styles_dict):
            dialog = PyStyleDialog(dialog_name=style, figure_window=figure_window)
            button = StaticSelectButton(style, f'pictures/icons/style_images/{style}.svg', style,
                                        f'pictures/icons/style_images/{style}.svg',
                                        dialog)
            self.button_dict[style] = button
            if index < 8:
                self.layout.addWidget(button)

        # 添加下拉按钮及其菜单
        self.pulldown_button = PullDownButton()
        self.stylemenu = StyleMenu(self.pulldown_button, self.button_dict)
        self.pulldown_button.connect_menu(self.stylemenu)

        self.layout.addWidget(self.pulldown_button)

        self.setLayout(self.layout)


class SelectorLayoutMenuBar(QFrame):
    def __init__(self, figure_window=None, fig_control_window=None):
        super().__init__()
        # 读取可用的样式
        style_json_path = os.path.join(current_path, 'available_layout.json')
        with open(style_json_path, 'r') as json_file:
            self.available_layout_dict = json.load(json_file)

        self.setObjectName("selector_menu")

        self.layout = QHBoxLayout()
        self.layout.setContentsMargins(0, 0, 0, 10)

        self.button_dict = {}

        for index, (layout, value) in enumerate(self.available_layout_dict.items()):
            dialog = PyLayoutDialog(dialog_name=layout, figure_window=figure_window, layout=value)
            button = StaticSelectButton(layout, f'pictures/icons/layout_images/{layout}.svg', layout,
                                        f'pictures/icons/layout_images/{layout}.svg', dialog)
            self.button_dict[layout] = button
            if index < 8:
                self.layout.addWidget(button)

        self.setLayout(self.layout)


class SelectorChartMenuBar(QFrame):
    """
    按钮链接的对话框由chart_dialog_dict提供
    """

    def __init__(self, figure_window=None):
        super().__init__()

        self.setObjectName("selector_menu")

        self.layout = QHBoxLayout()
        self.layout.setContentsMargins(0, 0, 0, 10)

        for index, (name, value) in enumerate(chart_dialog_dict.items()):
            # 传类进去，不传实例，点击按钮时才会创建实例，以便实时更新数据
            button = DynSelectButton(name, f'pictures/icons/chart_images/{name}.svg', name,
                                     f'pictures/icons/chart_images/{name}.svg', value, figure_window)
            if index < 8:
                self.layout.addWidget(button)

        self.setLayout(self.layout)


class SelectorElementMenuBar(QFrame):
    def __init__(self, figure_window=None):
        super().__init__()

        self.setObjectName("selector_menu")

        self.layout = QHBoxLayout()
        self.layout.setContentsMargins(0, 0, 0, 10)

        for index, (name, value) in enumerate(element_dialog_dict.items()):
            dialog = value(dialog_name=name, figure_window=figure_window)
            button = StaticSelectButton(name, f'pictures/icons/element_images/{name}.svg', name,
                                        f'pictures/icons/element_images/{name}.svg', dialog)
            if index < 8:
                self.layout.addWidget(button)

        self.setLayout(self.layout)
